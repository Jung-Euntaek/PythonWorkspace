from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져옴
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식 아닌 실제 데이터(열면서 계산하기 때문에 처음에는 None 출력, 새로 저장 한번 해야함)
for row in ws.values:
    for cell in row:
        print(cell)
