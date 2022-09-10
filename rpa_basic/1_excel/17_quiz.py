from openpyxl import load_workbook
wb = load_workbook("quiz.xlsx", data_only=True)
ws = wb.active

col_D = ws["D"]
for cell in col_D:
    if cell.row == 1:
        continue
    else:
        cell.value = 10

ws["H1"].value = "총점"
for i in range (2, 12):
    ws["H{}".format(i)] = "=SUM(B{}:G{})".format(i, i)



ws["I1"].value = "성적"

row_sum = 0
for row in ws.iter_rows(min_row=2):
    for i in range(1, 7):
        row_sum += int(row[i].value)
    if row_sum >= 90:
        row[8].value = "A"
    elif row_sum >= 80:
        row[8].value = "B"
    elif row_sum >= 70:
        row[8].value = "C"
    else:
        row[8].value = "D"
    row_sum = 0

    if int(row[1].value) < 5:
        row[8].value = "F"

wb.save("scores.xlsx")


# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active

# ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

# scores = [
# (1,10,8,5,14,26,12),
# (2,7,3,7,15,24,18),
# (3,9,5,8,8,12,4),
# (4,7,8,7,17,21,18),
# (5,7,8,7,16,25,15),
# (6,3,5,8,8,17,0),
# (7,4,9,10,16,27,18),
# (8,6,6,6,15,19,17),
# (9,10,10,9,19,30,19),
# (10,9,8,8,20,25,20)
# ]

# for s in scores:
#     ws.append(s)

# for idx, cell in enumerate(ws["D"]):
#     if idx == 0:
#         continue
#     cell.value = 10

# ws["H1"] = "총점"
# ws["I1"] = "성적"

# for idx, score in enumerate(scores, start=2): # 0-9까지가 아니라 2-11까지
#     sum_val = sum(score[1:]) - score[3] + 10
#     ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)

#     grade = None
#     if sum_val >= 90:
#         grade = "A"
#     elif sum_val >= 80:
#         grade = "B"
#     elif sum_val >= 70:
#         grade = "C"
#     else:
#         grade = "D"

#     if score[1] < 5:
#         grade = "F"

#     ws.cell(row=idx, column=9).value = grade

# wb.save("scores.xlsx")
