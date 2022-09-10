from os import W_OK
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 셀에 1 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3


ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀
print(ws["A1"].value) # A1 셀 값
print(ws["A10"].value) # 값이 없을 때는 None


ws.cell(row=1, column=1) # A1 셀
print(ws.cell(column=2, row=1).value)

ws.cell(column=3, row=1, value=10) # ws["C1"].value = 10 / ws["C1"] = 10

from random import *

# 반복문 이용해서 랜덤 숫자 채우기
idx = 1
for x in range(1, 11):
    for y in range(1, 11):
        # ws.cell(row=x, column=y, value=randint(1, 100)) # 행 먼저 입력
        ws.cell(row=x, column=y, value=idx)
        idx += 1


wb.save("sample.xlsx")
